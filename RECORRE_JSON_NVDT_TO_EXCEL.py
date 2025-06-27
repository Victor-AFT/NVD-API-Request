import os, json, csv
from datetime import datetime
from openpyxl import Workbook
import pandas as pd

from requests.exceptions import HTTPError
import requests

#REALIZAR PATH RELATIVOS
path_JSON='JSON\\'

now = datetime.now()
time = now.strftime("%m_%d_%Y_")

url='https://services.nvd.nist.gov/rest/json/cves/1.0/'
#name_json=path_JSON+time+"NVDT.json"
name_json='03_17_2022_NVDT.json'
file_name_csv=path_JSON+time+"format_NVDT.csv"

datos_cve=[]
lista_referencias=[]



def recoge_fabricante_cpe23uri(lista):
    separador=":"
    separado=lista.split(separador)
    return separado[3]

def recoge_firmware_cpe23uri(lista):
    separador=":"
    separado=lista.split(separador)
    return separado[4]

def get_http_and_exportJSON(url,NameJson):
    try:
        response = requests.get(url)
        response.raise_for_status()
        jsonResponse = response.json()
        file = open(NameJson, "w")
        json.dump(jsonResponse,file)
        file.close()
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Other error occurred: {err}')

if not os.path.exists(name_json):
    get_http_and_exportJSON(url,name_json)

f = open(name_json)
data = json.load(f)

"""variable para calcular la longitud"""
cve=data['result']['CVE_Items']

#myFile = open(file_name_csv, 'w', encoding='utf-8',newline='')
#writer = csv.writer(myFile)


fichero_excel=path_JSON+time+" CVES_NVDT.xlsx"


if not os.path.exists(fichero_excel):
    """CREA EL EXCEL CON LAS COLUMNAS Y LOS DATOS SI NO EXISTE"""

    # declaro funciones para crear el excel
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'ID_CVE'
    ws['B1'] = 'PUBLISHEDATE'
    #ws['C1'] = 'REFSOURCE'
    ws['C1'] = 'REFERENCES'
    ws['D1'] = 'DESCRIPTION'
    ws['E1'] = 'VECTORSTRING'
    ws['F1'] = 'BASESCORE'
    ws['G1'] = 'VERSION'
    ws['H1'] = 'VULNERABLE'
    ws['I1'] = 'FABRICANTE'
    ws['J1'] = 'FIRMWARE'
    ws['K1'] = 'CPE'
    ws['M1'] = 'VERSIONEND'

#Esta variable es para escribir apartir de la segunda linea del excel
    linea=2
    for cve_n in range(len(cve)):

        # print("Escribe en Linea: ",linea)
        # CVE 2022-0906
        cve_id = data['result']['CVE_Items'][cve_n]['cve']['CVE_data_meta']['ID']
        ws.cell(row=linea, column=1).value = cve_id
        # datos_cve.append(cve_id)

        # publishedDate
        cve_publishedDate = data['result']['CVE_Items'][cve_n]['publishedDate']
        ws.cell(row=linea, column=2).value = cve_publishedDate
        # datos_cve.append(cve_publishedDate)

        # REFERENCES URL
        references = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data']
        for r in range(len(references)):
            # references_refsource = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data'][r]['refsource']
            references_url = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data'][r]['url']
            # ws.cell(row=celda_referencias, column=3).value = references_refsource
            lista_referencias.append(references_url)
            # lista_referencias.append("--")
            ws.cell(row=linea, column=3).value = str(lista_referencias)
            # celda_referencias=celda_referencias+2

        # DESCRIPTION
        description = data['result']['CVE_Items'][cve_n]['cve']['description']['description_data'][0]['value']
        ws.cell(row=linea, column=4).value = description
        # datos_cve.append(description)

        # este campo puede estar vacio o no
        check_impact = data['result']['CVE_Items'][cve_n]['impact']
        if check_impact != {}:

            impact_baseMetricV3_vectorstring = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'vectorString']
            ws.cell(row=linea, column=5).value = impact_baseMetricV3_vectorstring
            # datos_cve.append(impact_baseMetricV3_vectorstring)

            impact_baseMetricV3_basescore = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'baseScore']
            ws.cell(row=linea, column=6).value = impact_baseMetricV3_basescore
            # datos_cve.append(impact_baseMetricV3_basescore)

            impact_baseMetricV3_version = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'version']
            ws.cell(row=linea, column=7).value = impact_baseMetricV3_version
            # datos_cve.append(impact_baseMetricV3_version)

        else:
            ws.cell(row=linea, column=5).value = 0
            ws.cell(row=linea, column=6).value = 0
            ws.cell(row=linea, column=7).value = 0

        check_configuration_nodes = data['result']['CVE_Items'][cve_n]['configurations']['nodes']
        if check_configuration_nodes != {}:

            for n in range(len(check_configuration_nodes)):
                # print("Escribe en Linea_cpe : ", linea)

                # HAY QUE MIRAR EL OPERATOR SI ES AND ES CHILDREN , SI ES OR ES CPE_MATCH
                # configuration_nodes_cpe_match_vulnerable_true_or_false=data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['vulnerable']
                configuration_nodes_operator = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n][
                    'operator']
                if configuration_nodes_operator == "OR":
                    configuration_nodes_cpe_match_vulnerable_true_or_false = \
                    data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['vulnerable']
                    # if configuration_nodes_cpe_match_vulnerable_true_or_false==True:
                    ws.cell(row=linea, column=8).value = configuration_nodes_cpe_match_vulnerable_true_or_false
                    # datos_cve.append(configuration_nodes_cpe_match_vulnerable_true_or_false)
                    configuration_nodes_cpe_match_vulnerable_true_cpe23uri = \
                    data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['cpe23Uri']

                    fabricante = recoge_fabricante_cpe23uri(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)
                    firmware = recoge_firmware_cpe23uri(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)

                    ws.cell(row=linea, column=9).value = fabricante
                    ws.cell(row=linea, column=10).value = firmware
                    ws.cell(row=linea, column=11).value = configuration_nodes_cpe_match_vulnerable_true_cpe23uri

                    datos_cve.append(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)
                    # datos_cve.append(fabricante)
                    # datos_cve.append(firmware)

                    configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding = \
                        data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]

                    if configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding in ['versionEndExcluding']:
                        # datos_cve.append(configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding)
                        ws.cell(row=linea,column=12).value = configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding


                    else:
                        ws.cell(row=linea, column=12).value = 0
                    linea = linea + 1

                else:

                    # OPERATOR AND
                    # CHILDREN PUEDE TENER 1 O VARIOS RESULTADOS
                    configuration_nodes_children = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n][
                        'children']

                    for c in range(len(configuration_nodes_children)):

                        configuration_nodes_children_cpe_match_vulnerable_true_or_false = \
                        data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c]['cpe_match'][0][
                            'vulnerable']

                        # if configuration_nodes_children_cpe_match_vulnerable_true_or_false == True:
                        ws.cell(row=linea,column=8).value = configuration_nodes_children_cpe_match_vulnerable_true_or_false
                        # datos_cve.append(configuration_nodes_children_cpe_match_vulnerable_true_or_false)
                        configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri = \
                            data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c][
                                'cpe_match'][0]['cpe23Uri']

                        # datos_cve.append(configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)

                        fabricante_children = recoge_fabricante_cpe23uri(
                            configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)
                        # datos_cve.append(fabricante_children)

                        firmware_children = recoge_firmware_cpe23uri(
                            configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)
                        # datos_cve.append(firmware_children)
                        ws.cell(row=linea, column=9).value = fabricante_children
                        ws.cell(row=linea, column=10).value = firmware_children
                        ws.cell(row=linea,column=11).value = configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri

                        configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding = \
                        data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c]['cpe_match'][0]

                        if configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding in [
                            'versionEndExcluding']:
                            # datos_cve.append(configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding)
                            ws.cell(row=linea,column=12).value = configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding
                        else:
                            ws.cell(row=linea, column=12).value = 0
                        linea = linea + 1
        else:
            linea = linea + 1

        lista_referencias = []

    wb.save(fichero_excel)



else:

    """lee la ultima fila que escribio y añade """

    #wb = Workbook()
    #ws = wb.active

    #xlsx=pd.read_excel(fichero_excel)
    #print(len(xlsx['ID_CVE']))
    #ultima_linea=len(xlsx['ID_CVE'])+1
    lista_datos=[]

    for cve_n in range(len(cve)):

        # CVE 2022-0906
        cve_id = data['result']['CVE_Items'][cve_n]['cve']['CVE_data_meta']['ID']
        #lista_datos.append(cve_id)
        lista_datos.append("111")

        # publishedDate
        cve_publishedDate = data['result']['CVE_Items'][cve_n]['publishedDate']
        lista_datos.append(cve_publishedDate)

        # REFERENCES URL
        references = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data']
        for r in range(len(references)):
            # references_refsource = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data'][r]['refsource']
            references_url = data['result']['CVE_Items'][cve_n]['cve']['references']['reference_data'][r]['url']
            # ws.cell(row=celda_referencias, column=3).value = references_refsource
            lista_referencias.append(references_url)

        lista_datos.append(str(lista_referencias))

        # DESCRIPTION
        description = data['result']['CVE_Items'][cve_n]['cve']['description']['description_data'][0]['value']
        lista_datos.append(description)

        # este campo puede estar vacio o no
        check_impact = data['result']['CVE_Items'][cve_n]['impact']
        if check_impact != {}:

            impact_baseMetricV3_vectorstring = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'vectorString']
            lista_datos.append(impact_baseMetricV3_vectorstring)

            impact_baseMetricV3_basescore = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'baseScore']
            lista_datos.append(impact_baseMetricV3_basescore)

            impact_baseMetricV3_version = data['result']['CVE_Items'][cve_n]['impact']['baseMetricV3']['cvssV3'][
                'version']
            lista_datos.append(impact_baseMetricV3_version)

        else:
            lista_datos.append('0')
            lista_datos.append('0')
            lista_datos.append('0')

        check_configuration_nodes = data['result']['CVE_Items'][cve_n]['configurations']['nodes']
        if check_configuration_nodes != {}:

            for n in range(len(check_configuration_nodes)):
                # print("Escribe en Linea_cpe : ", linea)

                # HAY QUE MIRAR EL OPERATOR SI ES AND ES CHILDREN , SI ES OR ES CPE_MATCH
                # configuration_nodes_cpe_match_vulnerable_true_or_false=data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['vulnerable']
                configuration_nodes_operator = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['operator']
                if configuration_nodes_operator == "OR":
                    configuration_nodes_cpe_match_vulnerable_true_or_false = \
                    data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['vulnerable']
                    # if configuration_nodes_cpe_match_vulnerable_true_or_false==True:


                    configuration_nodes_cpe_match_vulnerable_true_cpe23uri = \
                    data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]['cpe23Uri']

                    fabricante = recoge_fabricante_cpe23uri(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)
                    firmware = recoge_firmware_cpe23uri(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)




                    configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding = \
                        data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['cpe_match'][0]
                    if configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding in ['versionEndExcluding']:
                        # datos_cve.append(configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding)
                        lista_datos.append(configuration_nodes_cpe_match_vulnerable_true_versionEndExcluding)
                    else:
                       lista_datos.append('0')
                    #ultima_linea = ultima_linea + 1

                    lista_datos.append(configuration_nodes_cpe_match_vulnerable_true_or_false)
                    lista_datos.append(fabricante)
                    lista_datos.append(firmware)
                    lista_datos.append(configuration_nodes_cpe_match_vulnerable_true_cpe23uri)


                else:

                    # OPERATOR AND
                    # CHILDREN PUEDE TENER 1 O VARIOS RESULTADOS
                    configuration_nodes_children = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n][
                        'children']

                    for c in range(len(configuration_nodes_children)):

                        configuration_nodes_children_cpe_match_vulnerable_true_or_false = \
                        data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c]['cpe_match'][0]['vulnerable']


                        configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c]['cpe_match'][0]['cpe23Uri']



                        fabricante_children = recoge_fabricante_cpe23uri(configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)
                        firmware_children = recoge_firmware_cpe23uri(configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)


                        configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding = data['result']['CVE_Items'][cve_n]['configurations']['nodes'][n]['children'][c]['cpe_match'][0]

                        if configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding in [
                            'versionEndExcluding']:

                            lista_datos.append(configuration_nodes_children_cpe_match_vulnerable_true_versionEndExcluding)
                        else:
                            lista_datos.append('0')
                        #ultima_linea = ultima_linea + 1

                        lista_datos.append(configuration_nodes_children_cpe_match_vulnerable_true_or_false)
                        lista_datos.append(fabricante_children)
                        lista_datos.append(firmware_children)
                        lista_datos.append(configuration_nodes_children_cpe_match_vulnerable_true_cpe23uri)
        else:
            #ultima_linea = ultima_linea + 1

            lista_datos.append('0')
    df = pd.DataFrame([
                    [lista_datos[0], lista_datos[1],lista_datos[2],lista_datos[3],
                    lista_datos[4],lista_datos[5],lista_datos[6], lista_datos[7],
                    lista_datos[8],lista_datos[9], lista_datos[10],lista_datos[11]]],
        columns=['ID_CVE','PUBLISHEDATE','REFERENCES','DESCRIPTION','VECTORSTRING','BASESCORE','VERSION','VULNERABLE','FABRICANTE','FIRMWARE','CPE','VERSIONEND'])
    lista_datos=[]
    with pd.ExcelWriter(fichero_excel, mode="a", engine="openpyxl") as writer:
            df.to_excel(writer)

    lista_referencias = []







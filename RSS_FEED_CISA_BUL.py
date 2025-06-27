#!%userprofile%\AppData\Local\Programs\Python\Python38\python.exe

import feedparser, os, json, csv, re
from lxml.html import fromstring
from openpyxl import load_workbook
import time

cve_pattern = re.compile('CVE-\d{4}-\d{4,7}')
barraizq=re.compile('/')
guionmedio='-'
dobleguionmedio='--'

fichero_excel="test.xlsm"
wb = load_workbook(fichero_excel)
ws = wb.active
sheet_Inventario = wb['Inventario']

class CISAVulnerability:
    def __init__(self, vendor_product, description, published, cvss, cve_info):
        self.vendor_product = vendor_product
        self.description = description
        self.published = published
        self.cvss = cvss
        self.cve_info = cve_info

    def show_vulnerability_info(self):
        print("[*] Vulnerability found!")
        print("\t[-] Vendor-product: "+self.vendor_product)
        print("\t[-] Description: "+self.description)
        print("\t[-] Published: "+self.published)
        print("\t[-] Scoring: "+self.cvss)
        print("\t[-] CVE: " + self.cve_info)

strforma=""
fabricanteHoja=""
modeloHoja=""

vendorProduct=""
listcve=[]
def feed_vulnerability_reports():
    vulnerability_list = []
    rss = 'https://us-cert.cisa.gov/ncas/bulletins.xml'
    feed = feedparser.parse(rss)
    print('ULTIMO BOLETIN PUBLICADO: ',feed.entries[0]['title'])
    strforma = feed.entries[0]['title']
    strformat02 = strforma.replace(" ", "")
    strformat03 = strformat02.replace(",", "_")
    file_name_csv = str(strformat03) + '.csv'
    #print("aaa: ",feed.entries[0]['summary'])


    for key in feed["entries"]:
        published = key['published']
        summary = feed.entries[0]['summary']
        #print(summary)
        doc = fromstring(summary)
        tr_elements = doc.xpath('/tr')
        for j in range(1, len(tr_elements)):
            items = tr_elements[j]
            if len(items) != 5:
                break
            count = 0
            for t in items.iterchildren():
                data = t.text_content().strip()
                if count == 0:
                    vendor = data
                elif count == 1:
                    description = data
                elif count == 2:
                    published = data
                elif count == 3:
                    cvss = data
                elif count == 4:
                    info = data
                    cisa = CISAVulnerability(vendor, description, published, cvss, info)
                    #print(cisa)
                    vulnerability_list.append(cisa)
                count = count + 1

        if vulnerability_list is not None:
            for p in vulnerability_list:
                #p.show_vulnerability_info()
                print(p.vendor_product)
                stringcve = 'CVE-'
                cadenaencve = re.compile("[A-Za-z]")
                concatcve = ''
                for c in p.cve_info[4:]:
                    if not re.match(cadenaencve, c) and not re.match(barraizq, c):
                        concatcve += c
                stringcve += concatcve
                deltabulador = stringcve.replace('\t', "")
                delsaltline = deltabulador.replace('\n', "")
                # formateo pero recoger fabricante y modelo
                # print(p.vendor_product)
                cuentaguion = str(p.vendor_product).count('-')
                if int(cuentaguion) == 1:
                    # print(p.vendor_product)
                    concatunguion = str(p.vendor_product).replace("-", "--")
                    # print(concatunguion)
                else:
                    concatunguion = str(p.vendor_product)
                lista_fabricanteModelo = concatunguion.split("--")
                if len(lista_fabricanteModelo)>1:
                    fabricante = lista_fabricanteModelo[0]
                    modelo=lista_fabricanteModelo[1]
                    #print("FABICANTE CISA: ", fabricante, "-- MODEJO CISA: ", modelo)
                    """
                    for rows_hoja in sheet_Inventario.iter_rows(min_row=3, max_row=sheet_Inventario.max_row, min_col=1,max_col=2):
                        fabricanteHoja=str(rows_hoja[0].value)
                        fabricanteHojalower=str(fabricanteHoja.lower())
                        modeloHoja=str(rows_hoja[1].value)
                        modeloHojalower=modeloHoja.lower()
    
                        print("FABICANTE HOJA: ", fabricanteHojalower, "-- MODEJO HOJA: ", modeloHojalower)
                        print("FABICANTE CISA: ", fabricante, "-- MODEJO CISA: ", modelo)
                    """

feed_vulnerability_reports()




























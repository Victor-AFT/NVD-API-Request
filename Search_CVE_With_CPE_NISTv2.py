
import datetime
import os
import csv
from tqdm import tqdm
from openpyxl import load_workbook
import nvdlib
import requests

# Colores ANSI para la terminal
RESET = "\033[0m"
BOLD = "\033[1m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"

fichero_excel = 'Inventario_CPE.xlsx'
nombre_csv_datos = "vulnerabilidades.csv"

if not os.path.exists(fichero_excel):
    os.system("cls")
    print(f"{RED}Error: No existe el fichero {fichero_excel}{RESET}")
else:
    os.system("cls")

    print(f"{BOLD}#######################################{RESET}")
    print(f"{BOLD}############   API NIST  ###############{RESET}")
    print(f"{BOLD}###  Obtencion de CVE Con CPES v0.1  ###{RESET}")
    print(f"{BOLD}###  Mediante Fichero Excel          ###{RESET}")
    print(f"{BOLD}########################################\n{RESET}")

    numero_fila = 2
    wb = load_workbook(fichero_excel)
    inventario = wb['CPES']
    ws = wb.active

    fabmodver=[]
    cpe = []
    vulnerabilidad = []
    urls = []
    format_cvss = []
    buffer = []
    print(f"{YELLOW}El límite máximo de días para buscar vulnerabilidades publicadas o modificadas es de 120{RESET}")
    end = datetime.datetime.now()
    dia_introducido=int(input('Introduce el número de días a buscar: '))
    if dia_introducido > 120 or dia_introducido==0:
        print(f"{RED}El numero introducido debe ser superior a [1] o inferior a [120]{RESET}")
    else:
        if os.path.exists("key.conf"):
            f = open("key.conf", "r")
            line = f.readline()
            start = end - datetime.timedelta(days=int(dia_introducido))


            myfilecsv02 = open(nombre_csv_datos, 'w', encoding='utf-8', newline='')
            writer_r = csv.writer(myfilecsv02)

            print(f"Leyendo el archivo Excel {fichero_excel} y buscando vulnerabilidades... ")

            total_rows = inventario.max_row - 1
            progress_bar = tqdm(total=total_rows, desc="Procesando", unit="fila", bar_format="{l_bar}{bar}{r_bar}")
            detectada=0

            for row in inventario.iter_rows(min_row=2, max_row=inventario.max_row, min_col=1, max_col=inventario.max_column):
                fabricante = row[0].value
                modelo = row[1].value
                version = row[2].value
                cpes = row[3].value


                fabmodver.append(fabricante)
                fabmodver.append(modelo)
                fabmodver.append(version)

                for cpe in str(cpes).split(","):
                    try:
                        result = [r for r in nvdlib.searchCVE_V2(cpeName=cpe, key=str(line),pubStartDate=start, pubEndDate=end, delay=2)]
                        for rr in result:
                            vulnerabilidad.append(rr.id)
                            vulnerabilidad.append(rr.published)
                            vulnerabilidad.append(rr.lastModified)
                            solucion = [s for s in rr.references]
                            for url in solucion:
                                urls.append(url.url)
                            vulnerabilidad.append(urls)
                            urls = []
                            for des in rr.descriptions:
                                if des.lang == 'en':
                                    vulnerabilidad.append(des.value)

                            format_cvss.append(rr.score[0])

                            try:
                                format_cvss.append(rr.v31vector)
                            except:
                                pass
                            try:
                                format_cvss.append(rr.v3vector)
                            except:
                                pass
                            try:
                                format_cvss.append(rr.v2vector)
                            except:
                                pass

                            if len(format_cvss) == 2:
                                vulnerabilidad.append(format_cvss[0])
                            else:
                                vulnerabilidad.append(format_cvss[0])
                            format_cvss = []

                            vulnerabilidad.append(rr.score[1])
                            vulnerabilidad.append(rr.score[2])

                            if vulnerabilidad !=[]:
                                writer_r.writerow(fabmodver)
                                writer_r.writerow(vulnerabilidad)
                                #detectada=1


                            vulnerabilidad = []
                            os.system("cls")

                    except requests.exceptions.HTTPError as exc:
                        print(f"{RED}Error en cpe %s o key api no valida{RESET}"%(cpe))
                        os.system("cls")

                fabmodver=[]
                progress_bar.update(1)

            myfilecsv02.close()
            if os.path.getsize(nombre_csv_datos)>0:
                print(f"{GREEN}Se ha creado el archivo -> {nombre_csv_datos}{RESET}")
            else:
                try:
                    os.remove(nombre_csv_datos)
                    print(f"{GREEN} Vulnerabilidaes no detectadas.{RESET}")
                except FileNotFoundError:
                    pass
            progress_bar.close()

print("Programa finalizado")
os.system('pause')
